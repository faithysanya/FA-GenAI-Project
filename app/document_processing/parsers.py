"""
Document parsing module for extracting text from various file formats.
Supports: PDF, TXT, CSV, and Excel files.
"""

import os
from pathlib import Path
from typing import Optional
import chardet
import PyPDF2
import pandas as pd
import openpyxl


def parse_pdf(file_path: str) -> str:
    """
    Extract text from a PDF file using PyPDF2.
    
    Args:
        file_path: Path to the PDF file
        
    Returns:
        Extracted text from all pages of the PDF
        
    Raises:
        FileNotFoundError: If file doesn't exist
        Exception: If PDF parsing fails
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"PDF file not found: {file_path}")
    
    text = []
    try:
        with open(file_path, 'rb') as pdf_file:
            pdf_reader = PyPDF2.PdfReader(pdf_file)
            num_pages = len(pdf_reader.pages)
            
            for page_num in range(num_pages):
                page = pdf_reader.pages[page_num]
                text.append(page.extract_text())
        
        return '\n'.join(text)
    except Exception as e:
        raise Exception(f"Error parsing PDF {file_path}: {str(e)}")


def parse_txt(file_path: str) -> str:
    """
    Read text from a TXT file with automatic encoding detection.
    Tries UTF-8 first, then falls back to latin-1.
    
    Args:
        file_path: Path to the TXT file
        
    Returns:
        Text content of the file
        
    Raises:
        FileNotFoundError: If file doesn't exist
        Exception: If reading fails
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"TXT file not found: {file_path}")
    
    try:
        # Try to detect encoding
        with open(file_path, 'rb') as f:
            raw_data = f.read()
            detected = chardet.detect(raw_data)
            encoding = detected.get('encoding', 'utf-8') or 'utf-8'
        
        # Read with detected encoding
        try:
            with open(file_path, 'r', encoding=encoding) as f:
                return f.read()
        except (UnicodeDecodeError, LookupError):
            # Fall back to latin-1 if detection fails
            with open(file_path, 'r', encoding='latin-1') as f:
                return f.read()
    except Exception as e:
        raise Exception(f"Error parsing TXT {file_path}: {str(e)}")


def parse_csv(file_path: str) -> str:
    """
    Read CSV file and convert to formatted text using pandas.
    
    Args:
        file_path: Path to the CSV file
        
    Returns:
        Formatted text representation of CSV data
        
    Raises:
        FileNotFoundError: If file doesn't exist
        Exception: If parsing fails
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"CSV file not found: {file_path}")
    
    try:
        df = pd.read_csv(file_path)
        # Convert dataframe to formatted string with metadata
        output = f"CSV File: {Path(file_path).name}\n"
        output += f"Rows: {len(df)}, Columns: {len(df.columns)}\n"
        output += f"Columns: {', '.join(df.columns)}\n\n"
        output += df.to_string()
        return output
    except Exception as e:
        raise Exception(f"Error parsing CSV {file_path}: {str(e)}")


def parse_excel(file_path: str) -> str:
    """
    Extract text from an Excel file, including all sheets.
    Uses openpyxl to preserve formatting information.
    
    Args:
        file_path: Path to the Excel file (.xlsx or .xls)
        
    Returns:
        Formatted text from all sheets
        
    Raises:
        FileNotFoundError: If file doesn't exist
        Exception: If parsing fails
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found: {file_path}")
    
    try:
        output = []
        workbook = openpyxl.load_workbook(file_path)
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            output.append(f"\n{'='*50}")
            output.append(f"Sheet: {sheet_name}")
            output.append(f"{'='*50}\n")
            
            # Extract data from all cells
            for row in sheet.iter_rows(values_only=True):
                row_text = '\t'.join(str(cell) if cell is not None else '' for cell in row)
                output.append(row_text)
        
        return '\n'.join(output)
    except Exception as e:
        raise Exception(f"Error parsing Excel {file_path}: {str(e)}")


def parse_document(file_path: str, file_type: Optional[str] = None) -> str:
    """
    Router function that determines file type and calls appropriate parser.
    
    Args:
        file_path: Path to the document file
        file_type: Optional explicit file type ('pdf', 'txt', 'csv', 'excel').
                   If not provided, will be inferred from file extension.
    
    Returns:
        Extracted text from the document
        
    Raises:
        FileNotFoundError: If file doesn't exist
        ValueError: If file type is not supported
        Exception: If parsing fails
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")
    
    # Determine file type from extension if not provided
    if file_type is None:
        file_ext = Path(file_path).suffix.lower()
        file_type_map = {
            '.pdf': 'pdf',
            '.txt': 'txt',
            '.csv': 'csv',
            '.xlsx': 'excel',
            '.xls': 'excel',
        }
        file_type = file_type_map.get(file_ext)
        
        if file_type is None:
            raise ValueError(f"Unsupported file type: {file_ext}")
    
    # Route to appropriate parser
    file_type = file_type.lower().strip()
    
    if file_type == 'pdf':
        return parse_pdf(file_path)
    elif file_type == 'txt':
        return parse_txt(file_path)
    elif file_type == 'csv':
        return parse_csv(file_path)
    elif file_type in ['excel', 'xlsx', 'xls']:
        return parse_excel(file_path)
    else:
        raise ValueError(f"Unsupported file type: {file_type}")
